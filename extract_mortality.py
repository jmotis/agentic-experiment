"""
Extract parish burial and plague data from 1665.xlsx and 1666.xlsx
and combine into a single CSV: 1665-1666-burials-plague.csv
"""
import openpyxl
import csv
import re

MONTH_LOOKUP = {
    'jan': 'January', 'feb': 'February', 'mar': 'March',
    'apr': 'April',   'may': 'May',       'jun': 'June',
    'jul': 'July',    'aug': 'August',    'sep': 'September',
    'oct': 'October', 'nov': 'November',  'dec': 'December',
}
MONTH_NAMES = ['January','February','March','April','May','June',
               'July','August','September','October','November','December']


def normalize_month(s):
    """Turn any abbreviated/oddly-spelled month name into a canonical form."""
    clean = re.sub(r'[.\:\-]+', '', s).strip().lower()
    prefix = clean[:3]
    if prefix in MONTH_LOOKUP:
        return MONTH_LOOKUP[prefix]
    print(f"    WARNING: could not normalize month {s!r}")
    return s


def infer_from_sheet_name(sheet_name):
    """
    Fall-back: derive start/end date from the sheet tab name (MM_DD or MM-DD).
    Assumes a 7-day week.
    """
    parts = re.split(r'[-_]', sheet_name)
    if len(parts) != 2:
        return None
    try:
        mm, dd = int(parts[0]), int(parts[1])
        end_month = MONTH_NAMES[mm - 1]
        end_day   = str(dd)
        start_day_num = dd - 7
        if start_day_num <= 0:
            # Crosses a month boundary
            prev_mm = mm - 1 if mm > 1 else 12
            start_month = MONTH_NAMES[prev_mm - 1]
            days_in_prev = [31,28,31,30,31,30,31,31,30,31,30,31][prev_mm - 1]
            start_day = str(days_in_prev + start_day_num)
            final_end_month = end_month
        else:
            start_month = end_month
            start_day   = str(start_day_num)
            final_end_month = ''   # same month — leave End-Month blank
        print(f"    Inferred from sheet name '{sheet_name}': "
              f"{start_day} {start_month} → {end_day} {final_end_month or start_month}")
        return start_day, start_month, end_day, final_end_month
    except (ValueError, IndexError) as exc:
        print(f"    ERROR inferring from sheet name '{sheet_name}': {exc}")
        return None


def parse_header(header_text, sheet_name=None):
    """
    Parse the date string in the first cell of each sheet.
    Returns (start_day, start_month, end_day, end_month) where end_month
    is '' when the week stays within one calendar month.
    """
    if not header_text or not isinstance(header_text, str):
        return infer_from_sheet_name(sheet_name)

    # Match "From the START_DAY [of] START_MONTH to the REMAINDER"
    # Handles ordinal suffixes (1st, 2nd, 3rd, 6th) and optional "of"
    # (e.g. "From the 31 January ..." or "From the 1st of May ...").
    m = re.search(
        r'[Ff]rom\s+the\s*(\d+)(?:st|nd|rd|th|t|d)?\.?[^A-Za-z]*(?:of\s+)?'
        r'([A-Za-z][A-Za-z.\:\-]+?)\s+to\s+the\s+(.+)',
        header_text
    )
    if not m:
        print(f"    WARNING: header parse failed for {header_text!r}")
        return infer_from_sheet_name(sheet_name)

    start_day   = m.group(1)
    start_month = normalize_month(m.group(2))
    remainder   = m.group(3).strip()

    # Strip trailing year (4 digits) and any trailing punctuation/spaces
    remainder = re.sub(r'\s*\d{4}\.?\s*$', '', remainder).strip()
    # Strip editorial annotations like "[sic]" before processing
    remainder = re.sub(r'\[.*?\]', '', remainder).strip()

    # Extract end day, accounting for ordinal suffixes (st, nd, rd, th, t, d)
    end_m = re.match(r'(\d+)(?:st|nd|rd|th|t|d)?\.?[^A-Za-z]*', remainder)
    if not end_m:
        print(f"    WARNING: cannot extract end day from {remainder!r}")
        return infer_from_sheet_name(sheet_name)

    end_day    = end_m.group(1)
    after_day  = remainder[end_m.end():].strip()
    # Strip leading "of " if present
    after_day  = re.sub(r'^of\s+', '', after_day, flags=re.IGNORECASE).strip()

    # Whatever alphabetic token remains is the end month (if any)
    if after_day and re.match(r'^[A-Za-z]', after_day):
        tok = re.match(r'([A-Za-z][A-Za-z.\:\-]*)', after_day)
        end_month = normalize_month(tok.group(1)) if tok else ''
        if end_month == start_month:
            end_month = ''   # same month — leave End-Month blank in CSV
    else:
        end_month = ''

    return start_day, start_month, end_day, end_month


def is_skip_row(parish_name):
    """Return True for rows that should be excluded from output."""
    if not parish_name or not isinstance(parish_name, str):
        return True
    low = parish_name.lower().strip()
    # Summary totals
    if low.startswith('buried in'):
        return True
    # Christenings rows (any variety)
    if 'christn' in low or 'christening' in low:
        return True
    # Non-parish labels
    if low.startswith('page') or low.startswith('['):
        return True
    if low.startswith('source:') or 'eebo' in low or low.startswith('jmo'):
        return True
    if low.startswith('whereof'):
        return True
    # Diseases section header (loop stops here anyway, but just in case)
    if 'diseases and' in low and 'casual' in low:
        return True
    return False


def extract_sheet(ws, year, sheet_name):
    """Return a list of record dicts for one worksheet."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_text = rows[0][0] if rows[0] else None
    date_info   = parse_header(header_text, sheet_name)
    if date_info is None:
        print(f"    SKIPPING sheet '{sheet_name}' — cannot determine date")
        return []

    start_day, start_month, end_day, end_month = date_info
    records = []

    # rows[0] = date header, rows[1] = "Bur." / "Plag." column labels → skip both
    for row in rows[2:]:
        # Stop at "The Diseases and Casualties this Week."
        row_text = ' '.join(str(v) for v in row if v is not None)
        if 'diseases and' in row_text.lower() and 'casual' in row_text.lower():
            break

        # Three column triplets: cols 0-1-2, 3-4-5, 6-7-8
        for offset in (0, 3, 6):
            if offset + 2 >= len(row):
                continue
            parish = row[offset]
            bur    = row[offset + 1]
            plag   = row[offset + 2]

            if not isinstance(parish, str) or not parish.strip():
                continue
            if not isinstance(bur, (int, float)):
                continue
            if is_skip_row(parish):
                continue

            bur_val  = int(round(bur))
            plag_val = int(round(plag)) if isinstance(plag, (int, float)) else 0

            records.append({
                'Year':                  year,
                'Start-Day':             start_day,
                'Start-Month':           start_month,
                'End-Day':               end_day,
                'End-Month':             end_month,
                'Parish Name':           parish.strip(),
                'Total Parish Burials':  bur_val,
                'Parish Plague Burials': plag_val,
            })

    return records


def main():
    base = '/home/user/agentic-experiment'
    all_records = []

    for year_str in ('1665', '1666'):
        filepath = f'{base}/{year_str}.xlsx'
        print(f'\nProcessing {filepath}...')
        wb = openpyxl.load_workbook(filepath, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            recs = extract_sheet(ws, int(year_str), sheet_name)
            print(f'  {sheet_name}: {len(recs)} records')
            all_records.extend(recs)

    output = f'{base}/1665-1666-burials-plague.csv'
    fieldnames = [
        'Year', 'Start-Day', 'Start-Month', 'End-Day', 'End-Month',
        'Parish Name', 'Total Parish Burials', 'Parish Plague Burials',
    ]
    with open(output, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_records)

    print(f'\nDone — wrote {len(all_records)} records to {output}')


if __name__ == '__main__':
    main()
